"""One document in, plain text out, wrapped as untrusted.

`read_document` is one tool with a format router inside it, deliberately: the
grant surface stays a single checkbox and the model never has to work out which
file tool a file needs. Every lane ends in the same place — text, capped, inside
`<untrusted source="File …">`, with a note saying what was read and what was not.

The lanes, and why each is the cheapest honest reading of its format:

* **pdf** — pypdf's text layer first. A PDF that carries its own text needs no
  model at all, and a text layer is what the document itself says rather than
  what something inferred from a picture of it.
* **scans and images** — no text layer means the only way to read it is to look
  at it, so the document goes to the model on the extraction call path: no tools
  bound, the extraction system prompt's posture, the run's own model profile. The
  chat loop never carries the pixels, so a hostile page never shares a context
  with a bound tool. A profile that cannot see files refuses in one sentence.
* **xlsx / xls** — sheets are pages, rows and columns are bounded, and the cells
  are **values, never formulas**. A formula is code somebody wrote; the value is
  what a reader of the sheet sees.
* **csv, txt, md, json** — the bytes, decoded.
* **docx / pptx** — our own zipfile and xml.etree walk. Both formats are a zip of
  XML, and paragraphs, tables and slides are a hundred lines of it.
* **anything else** — refused by name, with what to save it as.

Nothing here writes anything, anywhere. Reading is not extraction: no draft, no
Document Extraction row, no sensitive-field decisions. The file is reached
through the same door every other file tool uses — `resolve_file` for the
reference, File's own read permission, then `require_provenance` for the anchor.
"""

import re
import zipfile
from io import BytesIO, StringIO
from typing import Any
from xml.etree import ElementTree

import filetype
import frappe
from frappe.utils import cint, flt

from frappe_agents.context.untrusted import wrap
from frappe_agents.extraction.pipeline import (
	DEFAULT_MAX_FILE_MB,
	pick_profile,
	read_source,
	require_provenance,
	resolve_file,
)
from frappe_agents.runner.providers import ExtractionRefused, ProviderError, call_model_extract
from frappe_agents.tools.base import run_model_profile

# What one call may return. Big enough for a real page of prose, small enough
# that a document cannot fill a context window on its own; past it the answer
# says which units it read, and the model asks for the next ones.
MAX_CHARS = 15_000

# Per sheet, per call. A spreadsheet is the one format where "the whole thing"
# is routinely millions of cells, so the bound is on the shape, not the bytes.
MAX_SHEET_ROWS = 200
MAX_SHEET_COLS = 30
MAX_CELL_CHARS = 500

# How many units one call may ask for. A range is convenience, not a way to ask
# for a thousand pages of work that the character cap will throw away anyway.
MAX_UNITS = 100

# The blank line between two units, counted against the cap like everything else.
SEPARATOR = "\n\n"
SEPARATOR_CHARS = len(SEPARATOR)

# Below this, a PDF's "text layer" is the stray glyph a scanner leaves behind and
# the document has to be looked at instead. Kept small on purpose: sending a page
# that really does carry its own text to a model costs money and needs a profile
# that can see files, so the doubt has to be genuine.
TEXT_LAYER_MIN_CHARS = 8

# A member of a docx or pptx is decompressed before it is parsed, so its own
# declared size is checked first: a zip bomb is 40 KB on disk.
MAX_XML_BYTES = 20 * 1024 * 1024

FORMAT_PDF = "pdf"
FORMAT_IMAGE = "image"
FORMAT_XLSX = "xlsx"
FORMAT_XLS = "xls"
FORMAT_DOCX = "docx"
FORMAT_PPTX = "pptx"
FORMAT_TEXT = "text"

LANE_TEXT_LAYER = "text layer"
LANE_VISION = "vision"
LANE_SPREADSHEET = "spreadsheet"
LANE_TEXT = "plain text"
LANE_DOCX = "docx xml"
LANE_PPTX = "pptx xml"

UNIT_PAGE = "page"
UNIT_SHEET = "sheet"
UNIT_SLIDE = "slide"
UNIT_PART = "part"
UNIT_DOCUMENT = "document"

READABLE = "PDF, PNG, JPEG, WEBP, GIF, XLSX, XLS, DOCX, PPTX, CSV, TXT, MD and JSON"

# The bytes are the truth about a file; the name is only a claim about it. So the
# sniffed type decides, and the extension is the fallback for the formats that
# have no signature to sniff — csv, txt, md and json are just text.
MIME_FORMATS = {
	"application/pdf": FORMAT_PDF,
	"image/png": FORMAT_IMAGE,
	"image/jpeg": FORMAT_IMAGE,
	"image/webp": FORMAT_IMAGE,
	"image/gif": FORMAT_IMAGE,
	"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": FORMAT_XLSX,
	"application/vnd.ms-excel": FORMAT_XLS,
	"application/vnd.openxmlformats-officedocument.wordprocessingml.document": FORMAT_DOCX,
	"application/vnd.openxmlformats-officedocument.presentationml.presentation": FORMAT_PPTX,
}

EXTENSION_FORMATS = {
	".pdf": FORMAT_PDF,
	".png": FORMAT_IMAGE,
	".jpg": FORMAT_IMAGE,
	".jpeg": FORMAT_IMAGE,
	".webp": FORMAT_IMAGE,
	".gif": FORMAT_IMAGE,
	".xlsx": FORMAT_XLSX,
	".xlsm": FORMAT_XLSX,
	".xls": FORMAT_XLS,
	".docx": FORMAT_DOCX,
	".pptx": FORMAT_PPTX,
	".csv": FORMAT_TEXT,
	".txt": FORMAT_TEXT,
	".md": FORMAT_TEXT,
	".markdown": FORMAT_TEXT,
	".json": FORMAT_TEXT,
}

# Formats a person will genuinely attach and this cannot read. Named back to them
# with what to do about it, because "unsupported file" is not an instruction.
REFUSED_MIMES = {
	"application/msword": "a legacy Word .doc",
	"application/vnd.ms-powerpoint": "a legacy PowerPoint .ppt",
	"application/vnd.oasis.opendocument.text": "an OpenDocument .odt",
	"application/vnd.oasis.opendocument.spreadsheet": "an OpenDocument .ods",
	"application/vnd.oasis.opendocument.presentation": "an OpenDocument .odp",
}
REFUSED_EXTENSIONS = {
	".doc": "a legacy Word .doc",
	".ppt": "a legacy PowerPoint .ppt",
	".odt": "an OpenDocument .odt",
	".ods": "an OpenDocument .ods",
	".odp": "an OpenDocument .odp",
	".rtf": "an RTF",
	".pages": "an Apple Pages",
	".numbers": "an Apple Numbers",
	".key": "an Apple Keynote",
}

# What the model is asked when a document has to be looked at. Fixed text: the
# tool takes a file and a page range and nothing else, so no wording a model
# writes can reach this call.
TRANSCRIPTION_INSTRUCTIONS = (
	"Transcribe this document. Write out every piece of text on it, in reading order, "
	"keeping tables as rows. Where something is a photograph, a stamp, a signature or a "
	"logo rather than text, describe it in one short phrase in square brackets. "
	"Do not summarise, do not explain, and do not add anything that is not on the document."
)
TRANSCRIPTION_SCHEMA = {
	"type": "object",
	"properties": {"text": {"type": "string", "description": "The document, transcribed."}},
	"required": ["text"],
	"additionalProperties": False,
}


def read_document(file_ref: str, pages: Any = None) -> dict:
	"""Read one attached document and return its text, wrapped as untrusted."""
	file_doc = resolve_file(file_ref)
	require_provenance(file_doc)

	settings = frappe.get_cached_doc("Agent Settings")
	content = _content(file_doc, settings)
	name = file_doc.file_name or file_doc.name
	wanted = parse_pages(pages)

	document_format = detect_format(content, name)
	if document_format == FORMAT_PDF:
		read = _read_pdf(file_doc, content, settings, wanted)
	elif document_format == FORMAT_IMAGE:
		read = _read_by_sight(file_doc, settings)
	elif document_format in (FORMAT_XLSX, FORMAT_XLS):
		read = _read_spreadsheet(document_format, content, wanted)
	elif document_format == FORMAT_DOCX:
		read = _assemble(_parts(docx_text(content)), wanted, UNIT_PART, LANE_DOCX)
	elif document_format == FORMAT_PPTX:
		read = _assemble(pptx_slides(content), wanted, UNIT_SLIDE, LANE_PPTX)
	else:
		read = _assemble(_parts(_decode(content)), wanted, UNIT_PART, LANE_TEXT)

	return _answer(file_doc, document_format, read)


def _answer(file_doc: Any, document_format: str, read: dict) -> dict:
	"""One reading, said the same way whichever lane produced it."""
	name = file_doc.file_name or file_doc.name
	text = read["text"]
	return {
		"file": name,
		"format": document_format,
		"lane": read["lane"],
		"unit": read["unit"],
		"read": read["read"],
		"total": read["total"],
		"truncated": read["truncated"],
		"note": read["note"],
		# The whole point of the wrapper: a document is data a stranger wrote, and
		# an instruction printed inside one is text to report, never to follow.
		"content": wrap(text, f"File {name}"),
		"_docs_touched": f"File: {file_doc.name}",
		"_result_summary": _summary(document_format, read, len(text)),
	}


def _summary(document_format: str, read: dict, chars: int) -> str:
	"""What the audit row says. Which lane ran is the fact worth keeping."""
	summary = f"{document_format} via {read['lane']}: {chars} chars"
	if read["total"]:
		summary += f", {read['unit']} {_span(read['read'])} of {read['total']}"
	if read.get("tokens"):
		summary += f", {read['tokens']} tokens"
	if read["truncated"]:
		summary += ", truncated"
	return summary


# --- getting at the bytes ---------------------------------------------------


def _content(file_doc: Any, settings: Any) -> bytes:
	"""The file's bytes, under the same size ceiling a document sent to a model has."""
	content = file_doc.get_content(encodings=[])
	if not isinstance(content, bytes):
		# get_content decodes anything that survives a text codec and the round trip
		# is lossy, which would turn a spreadsheet into plausible garbage.
		content = content.encode("utf-8", "surrogateescape")

	max_mb = cint(settings.get("max_extraction_file_mb")) or DEFAULT_MAX_FILE_MB
	size_mb = len(content) / (1024 * 1024)
	if size_mb > max_mb:
		# Measured on the bytes, never on File.file_size: that field is written from
		# the upload's own form data.
		raise ValueError(f"This file is {flt(size_mb, 1)} MB and reading is limited to {max_mb} MB.")
	return content


def detect_format(content: bytes, file_name: str) -> str:
	"""What this file is, sniffed first and named second."""
	kind = filetype.match(content)
	mime = (kind.mime if kind else "") or ""
	extension = _extension(file_name)

	if mime in MIME_FORMATS:
		return MIME_FORMATS[mime]
	if mime in REFUSED_MIMES:
		raise ValueError(_refusal(file_name, REFUSED_MIMES[mime]))
	if extension in EXTENSION_FORMATS:
		return EXTENSION_FORMATS[extension]
	if extension in REFUSED_EXTENSIONS:
		raise ValueError(_refusal(file_name, REFUSED_EXTENSIONS[extension]))

	named = f"a {extension.lstrip('.')} file" if extension else "of an unrecognised type"
	if mime:
		named = f"{named} ({mime})"
	raise ValueError(_refusal(file_name, named))


def _refusal(file_name: str, what: str) -> str:
	return (
		f"{file_name} is {what}, which cannot be read here. {READABLE} can be. "
		"Save it as PDF, DOCX or XLSX and attach it again."
	)


def _extension(file_name: str) -> str:
	name = (file_name or "").strip().lower()
	return f".{name.rsplit('.', 1)[1]}" if "." in name else ""


def _decode(content: bytes) -> str:
	"""Text from bytes, never failing: an undecodable byte is shown, not fatal."""
	return content.decode("utf-8-sig", errors="replace")


# --- which units were asked for ---------------------------------------------


def parse_pages(pages: Any) -> list[int] | None:
	"""`pages` as a list of unit numbers, or None for "as much as fits".

	One argument for every format: pages of a PDF, sheets of a spreadsheet,
	slides of a deck, and parts of a long flat file, which is what lets a model
	ask for the rest of anything with the same call it just made.
	"""
	if pages is None or pages == "" or pages == []:
		return None

	values: list[int] = []
	if isinstance(pages, bool):
		raise ValueError(_PAGES_HELP)
	elif isinstance(pages, int):
		values = [pages]
	elif isinstance(pages, list):
		for item in pages:
			values.extend(_page_range(str(item)))
	elif isinstance(pages, str):
		for part in pages.replace(" ", "").split(","):
			values.extend(_page_range(part))
	else:
		raise ValueError(_PAGES_HELP)

	wanted = sorted({value for value in values if value > 0})
	if not wanted:
		raise ValueError(_PAGES_HELP)
	if len(wanted) > MAX_UNITS:
		raise ValueError(f"That asks for {len(wanted)} at once; ask for at most {MAX_UNITS}.")
	return wanted


_PAGES_HELP = "pages must be a number or a range, like 2, '1-3' or '1,4-6'."


def _page_range(part: str) -> list[int]:
	part = part.strip()
	if not part:
		return []
	bounds = part.split("-")
	try:
		if len(bounds) == 1:
			return [int(bounds[0])]
		if len(bounds) == 2:
			first, last = int(bounds[0]), int(bounds[1])
			if last < first or last - first + 1 > MAX_UNITS:
				raise ValueError(_PAGES_HELP)
			return list(range(first, last + 1))
	except ValueError:
		raise ValueError(_PAGES_HELP)
	raise ValueError(_PAGES_HELP)


# --- putting the units together ---------------------------------------------


def _assemble(units: Any, wanted: list[int] | None, unit: str, lane: str, **extra: Any) -> dict:
	"""Take the units that were asked for, up to the cap, and say what was left.

	`units` is either a list of `(label, text)` pairs or a `(total, getter)` pair
	for a format where producing a unit costs something — a PDF page is parsed,
	not looked up, so a call for page 40 must not parse the other thirty-nine.
	"""
	total, get = units if isinstance(units, tuple) else (len(units), lambda number: units[number - 1])

	if not total:
		return dict(
			text="",
			lane=lane,
			unit=unit,
			read=[],
			total=0,
			truncated=False,
			note="There is nothing in this document to read.",
			**extra,
		)

	numbers = wanted or list(range(1, total + 1))
	beyond = [number for number in numbers if number > total]
	if beyond and len(beyond) == len(numbers):
		raise ValueError(f"This document has {total} {unit}(s), so there is no {unit} {beyond[0]}.")
	numbers = [number for number in numbers if number <= total]

	pieces: list[str] = []
	read: list[int] = []
	budget = MAX_CHARS
	cut = False
	for number in numbers:
		if budget <= 0:
			break
		label, text = get(number)
		# The cap is on what comes back, so what comes back is what it counts: the
		# unit's own heading and the blank line before it, not only its text. A
		# document of a thousand near-empty pages is otherwise unbounded — every
		# page costs nothing, so every page is read and every heading is returned.
		piece = f"[{label}]\n{text}" if total > 1 else text
		cost = len(piece) + (SEPARATOR_CHARS if pieces else 0)
		if cost > budget:
			if read:
				# Stop on a whole unit rather than half of one. The next call can
				# start here, and half a page read twice is worse than not read.
				break
			piece = piece[:budget]
			cost = len(piece)
			cut = True
		pieces.append(piece)
		read.append(number)
		budget -= cost

	truncated = cut or len(read) < len(numbers) or (not wanted and len(read) < total)
	return dict(
		text=SEPARATOR.join(pieces).strip(),
		lane=lane,
		unit=unit,
		read=read,
		total=total,
		truncated=truncated,
		note=_note(unit, read, total, numbers, cut),
		**extra,
	)


def _note(unit: str, read: list[int], total: int, numbers: list[int], cut: bool) -> str:
	"""One sentence about what came back, and how to ask for the rest."""
	if not read:
		return f"Nothing fitted in one call. Ask for one {unit} at a time."

	said = f"Read {unit} {_span(read)} of {total}."
	if cut:
		said += f" That {unit} was longer than one call holds and is cut off at the end."

	remaining = [number for number in range(1, total + 1) if number not in read]
	if not remaining:
		return said if total > 1 else "The whole document."
	if len(numbers) < total and read == numbers:
		# They asked for exactly this much and got it. Nothing to chase.
		return said
	return f"{said} Call read_document again with pages={_span(remaining[:MAX_UNITS])!r} for the rest."


def _span(numbers: list[int]) -> str:
	"""'1-4' when they run on, '1, 3, 7' when they do not."""
	if not numbers:
		return "none"
	if len(numbers) > 1 and numbers == list(range(numbers[0], numbers[-1] + 1)):
		return f"{numbers[0]}-{numbers[-1]}"
	return ", ".join(str(number) for number in numbers)


def _parts(text: str) -> list[tuple[str, str]]:
	"""A flat file cut into call-sized parts, on line boundaries where it can be.

	Formats without pages still need a way to be read to the end, so the cap
	itself becomes the unit and `pages` asks for the next part.
	"""
	text = text.strip()
	if not text:
		return []

	parts: list[tuple[str, str]] = []
	remaining = text
	while remaining:
		if len(remaining) <= MAX_CHARS:
			chunk, remaining = remaining, ""
		else:
			window = remaining[:MAX_CHARS]
			split = window.rfind("\n")
			cut = split + 1 if split > MAX_CHARS // 2 else MAX_CHARS
			chunk, remaining = remaining[:cut], remaining[cut:]
		parts.append((f"{UNIT_PART} {len(parts) + 1}", chunk.strip("\n")))
	return parts


# --- pdf --------------------------------------------------------------------


def _read_pdf(file_doc: Any, content: bytes, settings: Any, wanted: list[int] | None) -> dict:
	"""The text layer if there is one, and the model's eyes if there is not."""
	from pypdf import PdfReader
	from pypdf.errors import PdfReadError

	try:
		reader = PdfReader(BytesIO(content))
		# Asked before the page count, because on an encrypted file the page count
		# is itself a decryption failure and says so far less clearly.
		encrypted = reader.is_encrypted
	except PdfReadError as exc:
		raise ValueError(f"This PDF could not be read ({exc}).")
	if encrypted:
		raise ValueError("This PDF is password protected, so nothing in it can be read.")

	try:
		total = len(reader.pages)
	except PdfReadError as exc:
		raise ValueError(f"This PDF could not be read ({exc}).")

	def page(number: int) -> tuple[str, str]:
		try:
			text = reader.pages[number - 1].extract_text() or ""
		except Exception:
			# One unreadable page is a gap in the reading, not a failed read.
			text = ""
		return f"{UNIT_PAGE} {number}", text.strip()

	read = _assemble((total, page), wanted, UNIT_PAGE, LANE_TEXT_LAYER)
	if len(read["text"].strip()) >= TEXT_LAYER_MIN_CHARS:
		return read

	# No text layer worth the name: a scan, or pages that are pictures of text.
	looked = _read_by_sight(file_doc, settings)
	looked["note"] = f"This PDF has no text layer, so it was looked at rather than parsed. {looked['note']}"
	return looked


# --- scans and images -------------------------------------------------------


def _read_by_sight(file_doc: Any, settings: Any) -> dict:
	"""Transcribe a document by looking at it, on the extraction call path.

	Same call as an extraction: no tools bound, the extraction system prompt, the
	run's own model profile, and the file caps that path already enforces. What
	comes back is text, so the pixels stay out of the chat loop entirely — a
	hostile page is never in a context that holds a tool.
	"""
	source = read_source(file_doc, settings)
	profile = pick_profile(run_model_profile())

	try:
		reply = call_model_extract(
			profile,
			source["content"],
			source["media_type"],
			file_doc.file_name or "",
			TRANSCRIPTION_SCHEMA,
			TRANSCRIPTION_INSTRUCTIONS,
			schema_name="document_transcription",
		)
	except ExtractionRefused as exc:
		raise ValueError(f"The model declined to read this document. {exc}")
	except ProviderError as exc:
		# Includes the one refusal worth naming precisely: a profile that was never
		# declared able to see a file cannot read a scan, and says so.
		raise ValueError(str(exc))

	text = str((reply.get("data") or {}).get("text") or "").strip()
	if not text:
		raise ValueError(reply.get("parse_error") or "The model returned nothing for this document.")

	tokens = cint(reply.get("tokens_in")) + cint(reply.get("tokens_out"))
	return _assemble([(UNIT_DOCUMENT, text)], None, UNIT_DOCUMENT, LANE_VISION, tokens=tokens)


# --- spreadsheets -----------------------------------------------------------


def _read_spreadsheet(document_format: str, content: bytes, wanted: list[int] | None) -> dict:
	sheets = xlsx_sheets(content) if document_format == FORMAT_XLSX else xls_sheets(content)
	return _assemble(sheets, wanted, UNIT_SHEET, LANE_SPREADSHEET)


def xlsx_sheets(content: bytes) -> list[tuple[str, str]]:
	"""Every sheet as a bounded plain-text table. Values only — never formulas."""
	import openpyxl

	try:
		# data_only gives the value the sheet last computed. A formula is code
		# somebody wrote; what a reader of the sheet sees is the number.
		book = openpyxl.load_workbook(BytesIO(content), data_only=True, read_only=True)
	except Exception as exc:
		raise ValueError(f"This spreadsheet could not be read ({exc}).")

	sheets = []
	try:
		for index, title in enumerate(book.sheetnames, start=1):
			sheet = book[title]
			rows = [
				list(row)
				for row in sheet.iter_rows(
					max_row=MAX_SHEET_ROWS + 1, max_col=MAX_SHEET_COLS + 1, values_only=True
				)
			]
			sheets.append((f"{UNIT_SHEET} {index}: {title}", _table(rows)))
	finally:
		book.close()
	return sheets


def xls_sheets(content: bytes) -> list[tuple[str, str]]:
	"""The same table from a legacy .xls, read by xlrd."""
	import xlrd

	try:
		# xlrd narrates to stdout by default; a workbook missing a codepage record
		# is not news anybody wants in the bench log.
		book = xlrd.open_workbook(file_contents=content, logfile=StringIO())
	except Exception as exc:
		raise ValueError(f"This spreadsheet could not be read ({exc}).")

	sheets = []
	for index, sheet in enumerate(book.sheets(), start=1):
		rows = []
		for row in range(min(sheet.nrows, MAX_SHEET_ROWS + 1)):
			rows.append(
				[
					_xls_value(book, sheet.cell(row, column))
					for column in range(min(sheet.ncols, MAX_SHEET_COLS + 1))
				]
			)
		sheets.append((f"{UNIT_SHEET} {index}: {sheet.name}", _table(rows)))
	return sheets


def _xls_value(book: Any, cell: Any) -> Any:
	"""A legacy cell's value, with its dates turned back into dates."""
	import xlrd

	if cell.ctype == xlrd.XL_CELL_DATE:
		try:
			return xlrd.xldate.xldate_as_datetime(cell.value, book.datemode)
		except Exception:
			return cell.value
	if cell.ctype == xlrd.XL_CELL_BOOLEAN:
		return bool(cell.value)
	if cell.ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
		return None
	return cell.value


def _table(rows: list[list[Any]]) -> str:
	"""Rows as a plain-text table, bounded, saying what it left out."""
	extra_rows = len(rows) > MAX_SHEET_ROWS
	extra_columns = any(any(_cell(value) for value in row[MAX_SHEET_COLS:]) for row in rows)

	lines = []
	for row in rows[:MAX_SHEET_ROWS]:
		values = [_cell(value) for value in row[:MAX_SHEET_COLS]]
		while values and not values[-1]:
			values.pop()
		lines.append(" | ".join(values))
	while lines and not lines[-1]:
		lines.pop()

	notes = []
	if extra_rows:
		notes.append(f"only the first {MAX_SHEET_ROWS} rows are shown")
	if extra_columns:
		notes.append(f"only the first {MAX_SHEET_COLS} columns are shown")
	if not lines:
		# An empty sheet and a sheet of formulas nobody has ever calculated look
		# the same from here, and the difference matters to whoever asks.
		notes.append("no stored values: an empty sheet, or formulas that were never calculated")
	if notes:
		lines.append(f"({'; '.join(notes)})")
	return "\n".join(lines)


def _cell(value: Any) -> str:
	if value is None:
		return ""
	if isinstance(value, float) and value.is_integer():
		value = int(value)
	text = str(value).replace("\n", " ").strip()
	return text if len(text) <= MAX_CELL_CHARS else text[:MAX_CELL_CHARS] + "…"


# --- docx and pptx ----------------------------------------------------------

W = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
A = "{http://schemas.openxmlformats.org/drawingml/2006/main}"
SLIDE = re.compile(r"^ppt/slides/slide(\d+)\.xml$")


def docx_text(content: bytes) -> str:
	"""A Word document's paragraphs and tables, in the order they are written."""
	root = _xml_member(_office_zip(content, "Word document"), "word/document.xml", "Word document")
	body = root.find(f"{W}body")
	if body is None:
		return ""

	lines: list[str] = []
	for node in body:
		if node.tag == f"{W}p":
			lines.append(_docx_paragraph(node))
		elif node.tag == f"{W}tbl":
			lines.extend(_docx_table(node))
	return _tidy(lines)


def _docx_paragraph(node: Any) -> str:
	"""One paragraph: its runs joined, with tabs and breaks kept as themselves."""
	pieces = []
	for child in node.iter():
		if child.tag == f"{W}t":
			pieces.append(child.text or "")
		elif child.tag == f"{W}tab":
			pieces.append("\t")
		elif child.tag in (f"{W}br", f"{W}cr"):
			pieces.append("\n")
	return "".join(pieces).strip()


def _docx_table(node: Any) -> list[str]:
	"""A table as one line per row, cells separated the way a reader would."""
	lines = []
	for row in node.iter(f"{W}tr"):
		cells = [
			" ".join(part for part in (_docx_paragraph(p) for p in cell.iter(f"{W}p")) if part)
			for cell in row.iter(f"{W}tc")
		]
		lines.append(" | ".join(cells).strip())
	return lines


def pptx_slides(content: bytes) -> tuple[int, Any]:
	"""How many slides the deck has, and how to read any one of them.

	The lazy form, for the reason the PDF lane uses it: a deck is one small zip
	member per slide, so parsing every slide up front means opening the archive
	once per slide and reading its whole index each time. That is quadratic in
	the slide count — a two-megabyte deck of a few thousand slides costs minutes
	— and the cap throws nearly all of it away anyway.
	"""
	archive = _office_zip(content, "PowerPoint file")
	names = [name for name in archive.namelist() if SLIDE.match(name)]
	names.sort(key=lambda name: int(SLIDE.match(name).group(1)))

	def slide(number: int) -> tuple[str, str]:
		root = _xml_member(archive, names[number - 1], "PowerPoint file")
		lines = [
			"".join(node.text or "" for node in paragraph.iter(f"{A}t")) for paragraph in root.iter(f"{A}p")
		]
		return f"{UNIT_SLIDE} {number}", _tidy(lines)

	return len(names), slide


def _tidy(lines: list[str]) -> str:
	"""Drop the empty lines a document's markup leaves behind, keep the shape."""
	kept: list[str] = []
	for line in lines:
		line = line.rstrip()
		if line or (kept and kept[-1]):
			kept.append(line)
	return "\n".join(kept).strip()


def _office_zip(content: bytes, what: str) -> zipfile.ZipFile:
	"""An office file's archive, opened once. Its members are read from this one."""
	try:
		return zipfile.ZipFile(BytesIO(content))
	except zipfile.BadZipFile:
		raise ValueError(f"This {what} could not be opened.")


def _xml_member(archive: zipfile.ZipFile, member: str, what: str) -> Any:
	"""One XML part of an office file, opened defensively.

	Two things are checked before anything is parsed, because both are cheap to
	build and expensive to meet: the decompressed size of the member — a zip bomb
	is small on disk and enormous in memory — and any DTD, which is how an XML
	file asks a parser to expand itself or to go and fetch something.
	"""
	try:
		info = archive.getinfo(member)
	except KeyError:
		raise ValueError(f"This {what} has no {member} in it, so there is nothing to read.")
	if info.file_size > MAX_XML_BYTES:
		raise ValueError(f"This {what} is too large to read.")
	try:
		with archive.open(info) as handle:
			raw = handle.read(MAX_XML_BYTES + 1)
	except zipfile.BadZipFile:
		raise ValueError(f"This {what} could not be opened.")

	if len(raw) > MAX_XML_BYTES:
		raise ValueError(f"This {what} is too large to read.")
	if b"<!DOCTYPE" in raw or b"<!ENTITY" in raw:
		raise ValueError(f"This {what} declares its own XML entities, which are not read here.")

	try:
		return ElementTree.fromstring(raw)
	except ElementTree.ParseError as exc:
		raise ValueError(f"This {what} is damaged and could not be read ({exc}).")
